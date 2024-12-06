import os
import xml.etree.ElementTree as ET
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QProgressBar, QPushButton, QMessageBox
from openpyxl import load_workbook, Workbook
from qgis.core import QgsProject, QgsMessageLog, Qgis  # type: ignore
from .helper_functions import HelperBase

class GenerateExcelDialog(QDialog):
    
    def __init__(self, base_dir):
        super().__init__()
        self.helper = HelperBase()
        self.base_dir = base_dir
        self.template_path = 'templates'
        
        self.setWindowTitle("Generate Excel + XML Files")
        self.layout = QVBoxLayout()

        self.progress_bar = QProgressBar(self)
        self.layout.addWidget(self.progress_bar)

        self.run_button = QPushButton("Generate Excel + XML Files", self)
        self.run_button.clicked.connect(self.__exec__)
        self.layout.addWidget(self.run_button)

        self.setLayout(self.layout)

    def __exec__(self):
        layers = [
            QgsProject.instance().mapLayersByName('LINIE_JT')[0],
            QgsProject.instance().mapLayersByName('STALP_XML_')[0],
            QgsProject.instance().mapLayersByName('BRANSAMENT_XML_')[0],
            QgsProject.instance().mapLayersByName('GRUP_MASURA_XML_')[0],
            QgsProject.instance().mapLayersByName('FIRIDA_XML_')[0],
            QgsProject.instance().mapLayersByName('DESCHIDERI_XML_')[0],
            QgsProject.instance().mapLayersByName('TRONSON_predare_xml')[0]
        ]

        self.progress_bar.setMaximum(len(layers) * 2)  # Two steps per layer (XML and XLSX)
        self.progress_bar.setValue(0)
        self.generate_excel_xml(layers)

        # Notify user when all exports are complete
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle("Complete")
        msg_box.setText("File generation completed!")
        msg_box.setStandardButtons(QMessageBox.Ok)

        # Show the dialog and wait for the user's response
        if msg_box.exec_() == QMessageBox.Ok:
            self.close()  # Close the plugin dialog

    def generate_excel_xml(self, layers):
        """
        Generates XML and XLSX files for the columns of given layers, using predefined templates if available.
        Replaces blanks in column names with apostrophes.

        :param layers: List of QgsVectorLayer objects.
        """
        if not os.path.exists(self.base_dir):
            os.makedirs(self.base_dir)

        file_name_mapping = {
            "LINIE_JT": "linie_jt",
            "STALP_XML_": "stalp",
            "BRANSAMENT_XML_": "bransament",
            "GRUP_MASURA_XML_": "grup_masura",
            "FIRIDA_XML_": "firida",
            "DESCHIDERI_XML_": "deschidere",
            "TRONSON_predare_xml": "tronson_jt"
        }

        progress = 0
        for layer in layers:
            layer_name = layer.name()
            safe_layer_name = file_name_mapping.get(layer_name, layer_name)

            # Define paths
            xlsx_template_path = os.path.join(self.template_path, f"{safe_layer_name}.xlsx")
            xlsx_path = os.path.join(self.base_dir, f"{safe_layer_name}_.xlsx")
            xml_template_path = os.path.join(self.template_path, f"{safe_layer_name}.xml")
            xml_path = os.path.join(self.base_dir, f"{safe_layer_name}.xml")

            # Load or create Excel template
            if os.path.exists(xlsx_template_path):
                workbook = load_workbook(xlsx_template_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = safe_layer_name

                # Write header row
                headers = [field.name().replace(" ", "'") for field in layer.fields() if field.name().lower() != "fid"]
                sheet.append(headers)

            # Clear existing data rows (except the header)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None

            # Write data rows
            headers = [cell.value for cell in sheet[1]]  # Assuming the first row contains headers
            for i, feature in enumerate(layer.getFeatures(), start=2):  # Start from row 2
                for col, header in enumerate(headers, start=1):
                    sheet.cell(row=i, column=col, value=feature[header])

            # Save the populated Excel file
            workbook.save(xlsx_path)
            QgsMessageLog.logMessage(f"Saved populated template for '{layer_name}' as {xlsx_path}.", level=Qgis.Info)

            progress += 1
            self.progress_bar.setValue(progress)

            # Use XML template if available
            if os.path.exists(xml_template_path):
                self.populate_xml_template(xml_template_path, xml_path, layer)
            else:
                self.export_to_default_xml(xml_path, layer, safe_layer_name)
            
            progress += 1
            self.progress_bar.setValue(progress)

    def populate_xml_template(self, xml_template_path, xml_output_path, layer):
        """
        Populates an XML template with data from the given layer.
        
        :param xml_template_path: Path to the XML template file.
        :param xml_output_path: Path to save the populated XML file.
        :param layer: The QGIS vector layer containing the data.
        """
        tree = ET.parse(xml_template_path)
        root = tree.getroot()

        # Assuming the XML template has a repeating element that we need to populate with layer features
        repeating_element_tag = list(root)[0].tag  # Get the tag of the first repeating element
        parent = root

        # Remove existing entries (to refresh with new data)
        for child in root.findall(repeating_element_tag):
            parent.remove(child)

        # Populate with new data from the QGIS layer
        for feature in layer.getFeatures():
            new_element = ET.Element(repeating_element_tag)
            for field in layer.fields():
                field_name = field.name()
                field_value = feature[field_name]
                child_element = ET.SubElement(new_element, field_name)
                child_element.text = str(field_value) if field_value is not None else ""
            parent.append(new_element)

        # Write the updated XML to the output path
        tree.write(xml_output_path, encoding="utf-8-sig", xml_declaration=True)
        QgsMessageLog.logMessage(f"Populated XML template for '{layer.name()}' and saved to {xml_output_path}.", level=Qgis.Info)

    def export_to_default_xml(self, xml_output_path, layer, root_name):
        """
        Exports data to a default XML format if no template is available.
        
        :param xml_output_path: Path to save the XML file.
        :param layer: The QGIS vector layer containing the data.
        :param root_name: The name of the root XML element.
        """
        root = ET.Element(f"IGEA_{root_name.upper()}")
        
        for feature in layer.getFeatures():
            feature_elem = ET.SubElement(root, f"{root_name.upper()}_JT")
            for field in layer.fields():
                field_name = field.name()
                field_value = feature[field_name]
                field_elem = ET.SubElement(feature_elem, field_name)
                field_elem.text = str(field_value) if field_value is not None else ""
        
        tree = ET.ElementTree(root)
        tree.write(xml_output_path, encoding="utf-8-sig", xml_declaration=True)
        QgsMessageLog.logMessage(f"Exported default XML for '{layer.name()}' to {xml_output_path}.", level=Qgis.Info)
