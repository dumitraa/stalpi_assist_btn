from openpyxl import load_workbook

try:
    workbook = load_workbook("C:/Users/USER/AppData/Roaming/QGIS/QGIS3/profiles/default/python/plugins/stalpi_assist_buttons/func/templates/anexa.xlsx")
    print("Excel loaded successfully")
except Exception as e:
    print(f"Error loading Excel file: {e}")
    
    
try:
    workbook = load_workbook("C:/Users/USER/AppData/Roaming/QGIS/QGIS3/profiles/default/python/plugins/stalpi_assist_buttons/func/templates/anexa.xlsx")
    print("Excel loaded successfully2")
except Exception as e:
    print(f"Error loading Excel file2: {e}")