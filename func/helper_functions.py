import os
import xml.etree.ElementTree as ET
from xml.dom import minidom
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from qgis.core import QgsVectorLayer, QgsProject, QgsMessageLog, Qgis # type: ignore

from .parsers.firida import IgeaFiridaParser
from .parsers.bransament import IgeaBransamentParser
from .parsers.linie import IgeaLinieParser
from .parsers.tronson import IgeaTronsonParser
from .parsers.deschidere import IgeaDeschidereParser
from .parsers.stalp import IgeaStalpParser
from .parsers.grup_masura import IgeaGrupMasuraParser


class HelperBase:
    def __init__(self):
        super().__init__()
        
    # MARK: DEFAULT
    # Generate xml/xlsx files, replace blanks with apostrophes
    def generate_files(self, layers, base_dir):
        """
        Generates XML and XLSX files for the columns of given layers, skipping any column named 'fid'.
        Replaces blanks in column names with apostrophes.
        
        :param layers: List of QgsVectorLayer objects to process.
        :param base_dir: Directory where the files will be saved.
        """
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        
        for layer in layers:
            layer_name = layer.name()
            safe_layer_name = layer_name.replace(" ", "_")  # Replace spaces in layer name
            
            # XML file generation
            xml_path = os.path.join(base_dir, f"{safe_layer_name}.xml")
            root = ET.Element("Layer")
            root.set("name", layer_name)
            
            for field in layer.fields():
                if field.name().lower() != "fid":  # Skip 'fid'
                    field_elem = ET.SubElement(root, "Field")
                    field_name = field.name().replace(" ", "'")  # Replace blanks with apostrophes
                    field_elem.set("name", field_name)
                    field_elem.set("type", field.typeName())
            
            tree = ET.ElementTree(root)
            tree.write(xml_path, encoding="utf-8-sig", xml_declaration=True)
            
            # XLSX file generation
            xlsx_path = os.path.join(base_dir, f"{safe_layer_name}.xlsx")
            workbook = xlsxwriter.Workbook(xlsx_path)
            worksheet = workbook.add_worksheet()
            
            # Write header row
            headers = [field.name().replace(" ", "'") for field in layer.fields() if field.name().lower() != "fid"]
            for col_idx, header in enumerate(headers):
                worksheet.write(0, col_idx, header)
            
            # Write rows (data for illustration, can include actual feature data if needed)
            for row_idx, feature in enumerate(layer.getFeatures(), start=1):
                for col_idx, field in enumerate(layer.fields()):
                    if field.name().lower() != "fid":  # Skip 'fid'
                        value = feature[field.name()]
                        worksheet.write(row_idx, col_idx, value)
            
            workbook.close()
            
            print(f"Generated files for layer '{layer_name}':")
        
    
    # Retrieve layers by name from the QGIS project
    def get_layers(self):
        '''
        Get layers by name from the QGIS project and add them to self.layers
        '''
        QgsMessageLog.logMessage("Retrieving layers from the QGIS project...", "StalpiAssist", level=Qgis.Info)
        layers = {}
        layer_names = ['STALP_JT', 'TRONSON_JT', 'BRANS_FIRI_GRPM_JT', 'FB pe C LES', 'FIRIDA_RETEA_JT', 'GRID_GEIOD', 'PTCZ_PTAB', 'TRONSON_XML_', 'TRONSON_ARANJARE', 'poze', 'FIRIDA_XML_', 'BRANSAMENT_XML_', 'GRUP_MASURA_XML_', 'STALP_XML_', 'DESCHIDERI_XML_', 'TRONSON_predare_xml', 'LINIE_MACHETA', 'STALPI_MACHETA', 'TRONSON_MACHETA', 'FIRIDA MACHETA', 'GRUP MASURA MACHETA', 'DESCHIDERI MACHETA', 'BRANSAMENTE MACHETA', 'LINIE_JT']
        
        # Get all layers in the current QGIS project (keep the layer objects)
        qgis_layers = QgsProject.instance().mapLayers().values()
        QgsMessageLog.logMessage(f"----------- QGIS LAYERS: {qgis_layers}", "StalpiAssist", level=Qgis.Info)

        # Iterate through the actual layer objects
        for layer_name in layer_names:
            layer = next((l for l in qgis_layers if l.name() == layer_name), None)
            layers[layer_name] = layer  # Add the layer if found, else None
            # QgsMessageLog.logMessage(f"Layer found: key: {layer_name}, value: {layer}", "StalpiAssist", level=Qgis.Info)

        # QgsMessageLog.logMessage(f"Layers found with IDs: {layers}", "StalpiAssist", level=Qgis.Info)
        return layers


    def add_layer_to_project(self, layer_path):
        try:
            # Get the name of the layer without the file extension and the full path
            layer_name = os.path.splitext(os.path.basename(layer_path))[0]
            
            # Load the merged layer from the output path
            merged_layer = QgsVectorLayer(layer_path, layer_name, 'ogr')
            
            # Check if the layer is valid
            if not merged_layer.isValid():
                QgsMessageLog.logMessage(f"Invalid layer: {layer_path}", "StalpiAssist", level=Qgis.Critical)
                return
            
            # Add the layer to the project with the proper name
            QgsProject.instance().addMapLayer(merged_layer)
            # QgsMessageLog.logMessage(f"Layer added to project with name '{layer_name}': {layer_path}", "StalpiAssist", level=Qgis.Info)
            
        except Exception as e:
            QgsMessageLog.logMessage(f"Error adding layer to project: {e}", "StalpiAssist", level=Qgis.Critical)
            
            
    def run_algorithm(self, algorithm, params, context, feedback, outputs):
        try:
            results = algorithm.processAlgorithm(params, context, feedback)
            
            # Check if `outputs` is a list and validate all items
            if isinstance(outputs, list):
                found_all = True
                found_any = False
                
                for output in outputs:
                    if output in results and results[output]:
                        found_any = True
                    else:
                        found_all = False
                
                if found_all:
                    return True  # All outputs found and valid
                elif found_any:
                    return None  # Some outputs found but not all
                else:
                    return False  # No outputs found
            else:
                # Single output validation
                if outputs in results and results[outputs]:
                    return True
                return False
        except Exception as e:
            # Log the error for debugging
            print(f"Error running algorithm: {e}")
            return False

# MARK: PARSERS
    def save_xml(self, xml_name, name, xml_file):
        root = ET.Element(xml_name) 

        for linie in self.linii:
            linie_element = ET.SubElement(root, name)
            for attr, value in linie.__dict__.items():
                child = ET.SubElement(linie_element, attr.upper())
                child.text = str(value) if value is not None else ""

        rough_string = ET.tostring(root, 'utf-8')

        reparsed = minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="    ")

        with open(xml_file, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)
            
    def write_to_excel_sheet(self, parser, sheet_name, excel_file):
        data = []
        headers = list(self.mapping.keys())
        
        for element in parser:
            row = []
            for header in headers:
                mapping = self.mapping[header]
                if not mapping:
                    value = ""
                elif isinstance(mapping, tuple):
                    prefix, attr = mapping
                    value = f"{prefix} {getattr(element, attr, '')}"
                else:
                    value = getattr(element, mapping, "")
                row.append(value)
            data.append(row)
        
        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]
        
        start_row = 2
        existing_headers = {sheet.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1)}
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if header.strip(" ") in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value)
        
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for header in headers:
                if header in existing_headers:
                    cell = sheet.cell(row=row_idx, column=existing_headers[header])
                    cell.border = thin_border
        
        workbook.save(excel_file)
        
        
class SHPProcessor:
    '''
    Class to process the SHP layers, validate them and load them into QGIS
    '''
    def __init__(self, layers):
        '''
        Constructor for the SHPProcessor class
        :param layers: A dictionary with layer names and their respective QgsVectorLayer objects
        :param output_xlsx: The name of the output Excel file
        :return:
        '''
        self.layers = layers
        self.parsers = []
        self.invalid_elements = []
        self.load_layers()
    
    def load_layers(self):
        '''
        Load the SHP layers, parse them and store the parsers in a list
        :return: None
        '''
        for layer_name, layer in self.layers.items():
            match layer_name.lower():
                case "LINIE_JT": 
                    parser = IgeaLinieParser(layer)
                case "STALP_XML_":
                    parser = IgeaStalpParser(layer)
                case "BRANSAMENT_XML_":
                    parser = IgeaBransamentParser(layer)
                case "GRUP_MASURA_XML_":
                    parser = IgeaGrupMasuraParser(layer)
                case "DESCHIDERI_XML_":
                    parser = IgeaDeschidereParser(layer)
                case "FIRIDA_XML_":
                    parser = IgeaFiridaParser(layer)
                case "TRONSON_predare_xml":
                    parser = IgeaTronsonParser(layer)
                
            parser.parse()
            self.parsers.append(parser)
            QgsMessageLog.logMessage(f"Layer '{layer_name}' parsed successfully.", "StalpiAssist", level=Qgis.Info)
            QgsMessageLog.logMessage(f"Parser data: {parser.get_data()}", "StalpiAssist", level=Qgis.Info)
