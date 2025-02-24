from typing import List
from openpyxl import load_workbook
from qgis.core import QgsMessageLog, Qgis, QgsProject # type: ignore
from ... import config

class TronsonJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, denum, prop, class_id_loc, id_loc, nr_crt_loc, 
                 class_id_inc_tr, id_inc_tr, nr_crt_inc_tr, class_id_fin_tr, id_fin_tr, 
                 nr_crt_fin_tr, tip_tr, tip_cond, lung_tr, geo, sursa_coord, data_coord, 
                 unit_log_int, s_unit_log, post_luc, obs):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.denum = denum
        self.prop = prop
        self.class_id_loc = class_id_loc
        self.id_loc = id_loc
        self.nr_crt_loc = nr_crt_loc
        self.class_id_inc_tr = class_id_inc_tr
        self.id_inc_tr = id_inc_tr
        self.nr_crt_inc_tr = nr_crt_inc_tr
        self.class_id_fin_tr = class_id_fin_tr
        self.id_fin_tr = id_fin_tr
        self.nr_crt_fin_tr = nr_crt_fin_tr
        self.tip_tr = tip_tr
        self.tip_cond = tip_cond
        self.lung_tr = lung_tr
        self.geo = geo
        self.sursa_coord = sursa_coord
        self.data_coord = data_coord
        self.unit_log_int = unit_log_int
        self.s_unit_log = s_unit_log
        self.post_luc = post_luc
        self.obs = obs

    def __repr__(self):
        return f"TronsonJT(nr_crt={self.nr_crt}, denum={self.denum}, geo={self.geo})"


class IgeaTronsonParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.tronsoane: List[TronsonJT] = []
        
        self.mapping = {
            "Nr. crt": "nr_crt",
            "ID": "id_bdi",
            "Denumire": "denum",
            "Descrierea BDI": ("TR ", "denum"),
            "Proprietar": "prop",
            "ID_Locatia": "id_loc",
            "Locatia": lambda tr: self.get_linie_value(tr),
            "Nr.crt_Inceput de tronson": "nr_crt_inc_tr",
            "Inceput de tronson": lambda tr: tr.denum.split('-')[0].strip() if tr.denum else "",
            "Nr.crt_Final de tronson": "nr_crt_fin_tr",
            "Final de tronson": lambda tr: tr.denum.split('-')[1].strip() if tr.denum else "",
            "Tipul tronsonului": "tip_tr",
            "Tip conductor": "tip_cond",
            "Lungimea tronsonului (km)": "lung_tr",
            "Geometrie": "geo",
            "Sursa coordonate": "sursa_coord",
            "Data actualizarii coordonatelor": "data_coord",
            "Unitate logistica de intretinere": "unit_log_int",
            "Sectie unitate logistica": "s_unit_log",
            "Post de lucru": "post_luc",
            "Observatii": "obs"
        }
        
        # self.qgis_mapping = ["CLASS_ID", "ID_BDI", "NR_CRT", "DENUM", "PROP", "CLASS_ID_LOC", "ID_LOC", "NR_CRT_LOC", "CLASS_ID_INC_TR", "ID_INC_TR", "NR_CRT_INC_TR", "CLASS_ID_FIN_TR", "ID_FIN_TR", "NR_CRT_FIN_TR", "TIP_TR", "TIP_COND", "LUNG_TR", "GEO", "SURSA_COORD", "DATA_COORD", "UNIT_LOG_INT", "S_UNIT_LOG", "POST_LUC", "OBS"]
        

    def parse(self):
        try:
            if not self.vector_layer.isValid():
                raise ValueError("The provided layer is not valid.")
        except Exception as e:
            QgsMessageLog.logMessage(f"Error: {e}", "StalpiAssist", level=Qgis.Critical)
            return

        features = list(self.vector_layer.getFeatures())
        for feature in features:
            attributes = {key: feature[key] for key in feature.fields().names()}
            tronson_data = TronsonJT(
                id = feature.id(),
                class_id = attributes.get('CLASS_ID'),
                id_bdi = attributes.get('ID_BDI'),
                nr_crt = attributes.get('NR_CRT'),
                denum = attributes.get('DENUM'),
                prop = attributes.get('PROP'),
                class_id_loc = attributes.get('CLASS_ID_LOC'),
                id_loc = attributes.get('ID_LOC'),
                nr_crt_loc = attributes.get('NR_CRT_LOC'),
                class_id_inc_tr = attributes.get('CLASS_ID_INC_TR'),
                id_inc_tr = attributes.get('ID_INC_TR'),
                nr_crt_inc_tr = attributes.get('NR_CRT_INC_TR'),
                class_id_fin_tr = attributes.get('CLASS_ID_FIN_TR'),
                id_fin_tr = attributes.get('ID_FIN_TR'),
                nr_crt_fin_tr = attributes.get('NR_CRT_FIN_TR'),
                tip_tr = attributes.get('TIP_TR'),
                tip_cond = attributes.get('TIP_COND'),
                lung_tr = attributes.get('LUNG_TR'),
                geo = attributes.get('GEO'),
                sursa_coord = attributes.get('SURSA_COORD'),
                data_coord = attributes.get('DATA_COORD'),
                unit_log_int = attributes.get('UNIT_LOG_INT'),
                s_unit_log = attributes.get('S_UNIT_LOG'),
                post_luc = attributes.get('POST_LUC'),
                obs = attributes.get('OBS')
            )
            self.tronsoane.append(tronson_data)
            
    def get_name(self):
        return "TRONSON_predare_xml"

    def get_data(self):
        return self.tronsoane
    
    def resolve_mapping(self, parser, mapping):
        if isinstance(mapping, tuple):
            parts = [
                str(getattr(parser, element, "")).strip() if hasattr(parser, element) else str(element).strip()
                for element in mapping
            ]
            return " ".join(filter(None, parts)).strip()
        elif callable(mapping):
            return mapping(parser)
        return str(getattr(parser, mapping, "")).strip() if mapping else ""


    def write_to_excel_sheet(self, excel_file):
        data = []
        headers = list(self.mapping.keys())
        
        sorted_tr = sorted(
            self.tronsoane,
            key=lambda tr: tr.nr_crt if tr.nr_crt not in config.NULL_VALUES else float("inf")
        )
        for tronson in sorted_tr:
            row = []
            for header in headers:
                mapping = self.mapping.get(header)
                value = self.resolve_mapping(tronson, mapping)
                # Replace None or invalid values with an empty string
                value = "" if value in config.NULL_VALUES else value
                row.append(value)
            data.append(row)
        
        workbook = load_workbook(excel_file)
        sheet = workbook["TRONSON_JT"]
        
        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}
            
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if header.strip(" ") in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value if cell_value is not None else "")
        
        workbook.save(excel_file)
        
    def get_linie_value(self, feature):
        '''
        match with LINIE_JT ID_BDI and return LINIE_JT DENUM
        '''
        
        linie_layer = QgsProject.instance().mapLayersByName('LINIE_JT')[0]
        if not linie_layer:
            QgsMessageLog.logMessage("LINIE_JT layer not found.", "StalpiAssist", level=Qgis.Critical)
            return ""
        
        linie_features = linie_layer.getFeatures()
        matching_feature = [
            linie for linie in linie_features
            if linie['ID_BDI'] == feature.id_loc
        ]
        
        if not matching_feature:
            QgsMessageLog.logMessage("No matching feature found.", "StalpiAssist", level=Qgis.Warning)
            return ""
        
        return matching_feature[0]['DENUM'] if matching_feature else ""
