from typing import List
from openpyxl import load_workbook
from qgis.core import QgsMessageLog, Qgis, QgsProject # type: ignore
from ... import config
from ..helper_functions import HelperBase

class ConsolaJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, denum, class_id_loc, id_loc, nr_crt_loc, tip_cons, alt, sursa_coord, data_coord, geo, long, lat, x_stereo_70, y_stereo_70, z_stereo_70):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.denum = denum
        self.class_id_loc = class_id_loc
        self.id_loc = id_loc
        self.nr_crt_loc = nr_crt_loc
        self.tip_cons = tip_cons
        self.alt = alt
        self.sursa_coord = sursa_coord
        self.data_coord = data_coord
        self.geo = geo
        self.long = long
        self.lat = lat
        self.x_stereo_70 = x_stereo_70
        self.y_stereo_70 = y_stereo_70
        self.z_stereo_70 = z_stereo_70

    def __repr__(self):
        return f"ConsolaJT(denum={self.denum})"


class IgeaConsolaParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.console: List[ConsolaJT] = []
        self.helper = HelperBase()

        self.mapping = {
            "Nr.crt": "nr_crt",
            "Denumire": "denum", # NUME STR + NR. IMOBIL
            "Descrierea BDI": ("CONSOLA ", "denum"),
            "ID_Locatie": "id_loc",
            "Locatia": lambda cs: self.get_linie_value(cs),
            "ID_Descrierea instalatiei superioare": "class_id_inst_sup",
            "Descrierea instalatiei superioare": lambda cs: self.get_linie_value(cs),
            "Tip consola": "tip_cons", # ["de acoperis", "de zid"]
            "Latitudine (grade zecimale)": lambda cs: f"{float(cs.lat):.8f}" if cs.lat else "",
            "Longitudine (grade zecimale)": "long",
            "Altitudine (m)": "alt",
            "x - STEREO 70 (m)": "x_stereo_70",
            "y - STEREO 70 (m)": "y_stereo_70",
            "z - STEREO 70 (m)": "z_stereo_70",
            "Sursa coordonate": "sursa_coord",
            "Data actualizarii coordonatelor": "data_coord",
            "Geometrie": "geo", # POINT
        }
            
            
    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        for feature in self.vector_layer.getFeatures():
            stalp_data = ConsolaJT(
                id=feature.id(),
                class_id=feature["CLASS_ID"],
                id_bdi=feature["ID_BDI"],
                nr_crt=feature["NR_CRT"],
                denum=feature["DENUM"],
                class_id_loc=feature["CLASS_ID_LOC"],
                id_loc=feature["ID_LOC"],
                nr_crt_loc=feature["NR_CRT_LOC"],
                tip_cons=feature["TIP_CONS"],
                alt=feature["ALT"],
                sursa_coord=feature["SURSA_COORD"],
                data_coord=feature["DATA_COORD"],
                geo=feature["GEO"],
                long=feature["LONG"],
                lat=feature["LAT"],
                x_stereo_70=feature["X_STEREO_70"],
                y_stereo_70=feature["Y_STEREO_70"],
                z_stereo_70=feature["Z_STEREO_70"]
            )
            self.console.append(stalp_data)
            
    def get_name(self):
        return "CONSOLA_XML_"
            
    def get_data(self):
        return self.console

    def get_linie_value(self, feature):
        '''
        match with LINIE_JT ID_BDI and return LINIE_JT DENUM
        '''
        
        linie_layer = QgsProject.instance().mapLayersByName('LINIE_JT')[0]
        if not linie_layer:
            QgsMessageLog.logMessage("LINIE_JT layer not found.", "StalpiAssist", level=Qgis.Critical)
            return ""
        
        linie_features = linie_layer.getFeatures()
        matching_feature = [
            linie for linie in linie_features
            if linie['ID_BDI'] == feature.id_inst_sup
        ]
        
        if not matching_feature:
            QgsMessageLog.logMessage("No matching feature found.", "StalpiAssist", level=Qgis.Warning)
            return ""
        
        return matching_feature[0]['DENUM'] if matching_feature else ""

    
    def write_to_excel_sheet(self, excel_file):
        data = []
        headers = list(self.mapping.keys())
        
        sorted_st = sorted(
            self.console,
            key=lambda st: st.nr_crt if st.nr_crt not in config.NULL_VALUES else float("inf")
        )
        for stalp in sorted_st:
            row = []
            for header in headers:
                mapping = self.mapping.get(header)
                value = self.helper.resolve_mapping(stalp, mapping)
                value = "" if value in config.NULL_VALUES else value
                row.append(value)
            data.append(row)
        
        workbook = load_workbook(excel_file)
        sheet = workbook["CONSOLA_PE_CLADIRE"]
        
        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if self.helper.n(header) in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value if cell_value is not None else "")

        workbook.save(excel_file)

