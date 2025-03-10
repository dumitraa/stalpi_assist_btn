from typing import List
from openpyxl import load_workbook
import pandas as pd
from qgis.core import QgsMessageLog, Qgis, QgsProject # type: ignore
from ... import config
from ..helper_functions import HelperBase

class FiridaJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, iden, class_id_loc, id_loc, nr_crt_loc, 
                 class_id_inst_sup, id_inst_sup, nr_crt_inst_sup, jud, prim, loc, tip_str, 
                 str, nr, etaj, rol_firi, tip_firi_ret, tip_firi_br, ampl, mat, lim_prop, 
                 def_firi, nr_cir, an_func, alt, geo, sursa_coord, data_coord, long, lat, 
                 x_stereo_70, y_stereo_70, z_stereo_70):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.iden = iden
        self.class_id_loc = class_id_loc
        self.id_loc = id_loc
        self.nr_crt_loc = nr_crt_loc
        self.class_id_inst_sup = class_id_inst_sup
        self.id_inst_sup = id_inst_sup
        self.nr_crt_inst_sup = nr_crt_inst_sup
        self.jud = jud
        self.prim = prim
        self.loc = loc
        self.tip_str = tip_str
        self.str = str
        self.nr = nr
        self.etaj = etaj
        self.rol_firi = rol_firi
        self.tip_firi_ret = tip_firi_ret
        self.tip_firi_br = tip_firi_br
        self.ampl = ampl
        self.mat = mat
        self.lim_prop = lim_prop
        self.def_firi = def_firi
        self.nr_cir = nr_cir
        self.an_func = an_func
        self.alt = alt
        self.geo = geo
        self.sursa_coord = sursa_coord
        self.data_coord = data_coord
        self.long = long
        self.lat = lat
        self.x_stereo_70 = x_stereo_70
        self.y_stereo_70 = y_stereo_70
        self.z_stereo_70 = z_stereo_70

    def __repr__(self):
        return f"FiridaJT(nr_crt={self.nr_crt}, iden={self.iden}, geo={self.geo})"


class IgeaFiridaParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.firide: List[FiridaJT] = []
        self.helper = HelperBase()
        
        self.mapping = {
            "Nr.crt": "nr_crt",
            "Identificator": lambda fr: fr.iden['correct'],
            "Descrierea BDI": lambda fr: "FR " if fr.rol_firi == "de retea" else "FB " + fr.iden['correct'],
            "ID_Locatia": "nr_crt_loc",
            "Locatia": lambda fr: "BR " + fr.iden['correct'],
            "ID_Descrierea instalatiei superioare": "id_inst_sup",
            "Descrierea instalatiei superioare": lambda fr: self.get_linie_value(fr),
            "Judet": "jud",
            "Primarie": "prim",
            "Localitate": "loc",
            "Tip strada": "tip_str",
            "Strada": lambda fr: fr.iden['first_str'],
            "Numarul": lambda fr: fr.iden['nr'],
            "Etaj": "etaj",
            "Rolul firidei": "rol_firi",
            "Tip firida retea": "tip_firi_ret",
            "Amplasare": "ampl",
            "Material": "mat",
            "Defecte firida": "def_firi",
            "Nr circuite": "nr_cir",
            "Tip firida bransament": "tip_firi_br",
            "Limita de proprietate": "lim_prop",
            "Anul punerii în functiune": "an_func",
            "Latitudine (grade zecimale)": "lat",
            "Longitudine (grade zecimale)": "long",
            "Altitudine (m)": "alt",
            "x - STEREO 70 (m)": "x_stereo_70",
            "y - STEREO 70 (m)": "y_stereo_70",
            "z - STEREO 70 (m)": "z_stereo_70",
            "Geometrie": "geo",
            "Sursa coordonate": "sursa_coord",
            "Data actualizarii coordonatelor": "data_coord"
        }

    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        features = list(self.vector_layer.getFeatures())
        for feature in features:
            attributes = {key: feature[key] for key in feature.fields().names()}
            fr_iden = self.helper.get_fr_iden(feature, True)
            firida_data = FiridaJT(
                id = feature.id(),
                class_id = attributes.get('CLASS_ID'),
                id_bdi = attributes.get('ID_BDI'),
                nr_crt = attributes.get('NR_CRT'),
                iden = fr_iden,
                class_id_loc = attributes.get('CLASS_ID_LOC'),
                id_loc = attributes.get('ID_LOC'),
                nr_crt_loc = attributes.get('NR_CRT_LOC'),
                class_id_inst_sup = attributes.get('CLASS_ID_LOC'),
                id_inst_sup = attributes.get('ID_INST_SUP'),
                nr_crt_inst_sup = attributes.get('NR_CRT_INST_SUP'),
                jud = attributes.get('JUD'),
                prim = attributes.get('PRIM'),
                loc = attributes.get('LOC'),
                tip_str = attributes.get('TIP_STR'),
                str = attributes.get('STR'),
                nr = attributes.get('NR'),
                etaj = attributes.get('ETAJ'),
                rol_firi = attributes.get('ROL_FIRI'),
                tip_firi_ret = attributes.get('TIP_FIRI_RET'),
                tip_firi_br = attributes.get('TIP_FIRI_BR'),
                ampl = attributes.get('AMPL'),
                mat = attributes.get('MAT'),
                lim_prop = attributes.get('LIM_PROP'),
                def_firi = attributes.get('DEF_FIRI'),
                nr_cir = attributes.get('NR_CIR'),
                an_func = attributes.get('AN_FUNC'),
                alt = attributes.get('ALT'),
                geo = attributes.get('GEO'),
                sursa_coord = attributes.get('SURSA_COORD'),
                data_coord = attributes.get('DATA_COORD'),
                long = attributes.get('LONG'),
                lat = attributes.get('LAT'),
                x_stereo_70 = attributes.get('X_STEREO_70'),
                y_stereo_70 = attributes.get('Y_STEREO_70'),
                z_stereo_70 = attributes.get('Z_STEREO_70')
            )
            self.firide.append(firida_data)

    def get_name(self):
        return "FIRIDA_XML_"

    def get_data(self):
        return self.firide

    def write_to_excel_sheet(self, excel_file, split=False, done_split=False):
        data = []
        headers = list(self.mapping.keys())

        # Collect data rows
        sorted_fr = sorted(
            self.firide,
            key=lambda fr: fr.nr_crt if fr.nr_crt not in config.NULL_VALUES else float("inf")
        )
        for firida in sorted_fr:
            row = []
            for header in headers:
                mapping = self.mapping[header]
                value = self.helper.resolve_mapping(firida, mapping)
                value = "" if value in config.NULL_VALUES else value
                row.append(value)
            data.append(row)

        df = pd.DataFrame(data, columns=headers)

        # Open Excel file
        try:
            workbook = load_workbook(excel_file)
        except Exception as e:
            QgsMessageLog.logMessage(f"Error opening workbook: {e}", "StalpiAssist", level=Qgis.Critical)
            return

        if split:
            df_retea = df[df['Rolul firidei'].str.contains('retea', na=False)]
            sheets = {'FIRIDA RETEA': df_retea}
        else:
            sheets = {'FIRIDA': df}

        for sheet_name, df_sheet in sheets.items():
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)

            # Determine header and start rows
            start_row = sheet.max_row + 1
            header_row = sheet.max_row - 1
            existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}

            # Ensure headers exist
            header_mapping = {header: existing_headers.get(header, None) for header in headers}


            # Write data rows
            for row_idx, row_data in enumerate(df_sheet.itertuples(index=False, name=None), start=start_row):
                for col_name, cell_value in zip(headers, row_data):
                    col_idx = header_mapping.get(col_name)
                    if col_idx:  # Only write to columns that exist
                        sheet.cell(row=row_idx, column=col_idx, value=cell_value if cell_value is not None else "")
                    
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=header_row, column=col_idx)
                if cell.value == "Nr circuite":
                    cell.value = "Nr circuite "
                    break

        try:
            workbook.save(excel_file)
        except Exception as e:
            QgsMessageLog.logMessage(f"Error saving workbook: {e}", "StalpiAssist", level=Qgis.Critical)

        if not done_split:
            self.write_to_excel_sheet(excel_file, split=True, done_split=True)

    def get_linie_value(self, feature):
        '''
        match with LINIE_JT ID_BDI and return LINIE_JT DENUM
        '''
        
        linie_layer = QgsProject.instance().mapLayersByName('LINIE_JT')[0]
        if not linie_layer:
            QgsMessageLog.logMessage("LINIE_JT layer not found.", "StalpiAssist", level=Qgis.Critical)
            return ""
        
        linie_features = linie_layer.getFeatures()
        matching_feature = [
            linie for linie in linie_features
            if linie['ID_BDI'] == feature.id_inst_sup
        ]
        
        if not matching_feature:
            QgsMessageLog.logMessage("No matching feature found.", "StalpiAssist", level=Qgis.Warning)
            return ""
        
        return matching_feature[0]['DENUM'] if matching_feature else ""