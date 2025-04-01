from typing import List
from openpyxl import load_workbook
from ... import config
from ..helper_functions import HelperBase
from qgis.core import QgsMessageLog, Qgis # type: ignore
from qgis.PyQt.QtWidgets import QMessageBox # type: ignore
import os
import pandas as pd
from pathlib import Path

class LinieJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, denum, prop, class_id_loc, id_loc, class_id_inst_sup, desc_inst_sup, cod_ad_energ, niv_ten, tip_lin, an_pif_init, nr_iv):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.denum = denum
        self.prop = prop
        self.class_id_loc = class_id_loc
        self.id_loc = id_loc
        self.class_id_inst_sup = class_id_inst_sup
        self.desc_inst_sup = desc_inst_sup
        self.cod_ad_energ = cod_ad_energ
        self.niv_ten = niv_ten
        self.tip_lin = tip_lin
        self.an_pif_init = an_pif_init
        self.nr_iv = nr_iv

    def __repr__(self):
        return f"LinieJT(denum={self.denum}, niv_ten={self.niv_ten}, tip_lin={self.tip_lin})"


class IgeaLinieParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.linii: List[LinieJT] = []
        self.helper = HelperBase()
        
        self.mapping = {
            "ID": "id_bdi",
            "Denumire": lambda _: "",
            "Descrierea BDI": "denum",
            "Proprietar": lambda ln: ln.prop if ln.prop not in config.NULL_VALUES else "DEER",
            "Locatia": "id_loc",
            "Descrierea instalatiei superioare": "desc_inst_sup",
            "Nivel tensiune (kV)": "niv_ten",
            "Tipul liniei": "tip_lin",
        }
        
        # self.qgis_mapping = ["CLASS_ID", "ID_BDI", "NR_CRT", "DENUM", "PROP", "CLASS_ID_LOC", "ID_LOC", "CLASS_ID_INST_SUP", "ID_INST_SUP", "COD_AD_ENERG", "NIV_TEN", "TIP_LIN", "AN_PIF_INIT", "NR_IV"]

    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        features = list(self.vector_layer.getFeatures())
        for feature in features:
            attributes = {key: feature[key] for key in feature.fields().names()}
            linie_data = LinieJT(
                id=feature.id(),
                class_id=attributes.get("CLASS_ID"),
                id_bdi=attributes.get("ID_BDI"),
                nr_crt=attributes.get("NR_CRT"),
                denum=attributes.get("DENUM"),
                prop=attributes.get("PROP"),
                class_id_loc=attributes.get("CLASS_ID_LOC"),
                id_loc=attributes.get("ID_LOC"),
                class_id_inst_sup=attributes.get("CLASS_ID_INST_SUP"),
                desc_inst_sup=self.get_descriere(feature),
                cod_ad_energ=attributes.get("COD_AD_ENERG"),
                niv_ten='0,4',
                tip_lin='LEA JT',
                an_pif_init=attributes.get("AN_PIF_INIT"),
                nr_iv=attributes.get("NR_IV")
            )
            self.linii.append(linie_data)
                    
    def get_name(self):
        return "LINIE_JT"
            
    def get_data(self):
        return self.linii

    def load_lookup_values(self, xlsx_path):
        """
        Loads values from the given Excel file and returns a set of lookup values.
        """
        if not os.path.exists(xlsx_path):
            QgsMessageLog.logMessage(f"File not found: {xlsx_path}", "StalpiAssist", level=Qgis.Critical)
            return set()

        try:
            df = pd.read_excel(xlsx_path, usecols=[0], dtype=str)
            return set(df.iloc[:, 0].dropna().str.strip())
        except Exception as e:
            QgsMessageLog.logMessage(f"Error loading Excel file: {e}", "StalpiAssist", level=Qgis.Critical)
            return set()

    @staticmethod
    def plugin_path(*args) -> Path:
        """ Return the path to the plugin root folder or file. """
        path = Path(__file__).resolve().parent
        for item in args:
            path = path.joinpath(item)
        return path

    def get_descriere(self, feature):
        """
        Extracts the matching substring from feature["DENUM"] if it exists in the lookup values.
        """
        denum_value = str(feature["DENUM"]).strip().replace("_", " ")
        if not denum_value:
            return None

        # Ensure lookup values are loaded once and cached
        if not hasattr(self, '_lookup_values'):
            xlsx_path = self.plugin_path('..', 'templates', 'pt.xlsx')
            self._lookup_values = self.load_lookup_values(xlsx_path)

        # Check for matches in the lookup set
        for lookup in self._lookup_values:
            if lookup.replace("_", " ") in denum_value:
                return lookup  # Return first found match

        QMessageBox.critical(None, "Avertizare", f"Nu a fost gasit niciun match pentru valoarea '{denum_value}'. S-a folosit denumirea proiectului")
        project_name = self.helper.get_pt_name()
        
        return project_name  # No match found

    def write_to_excel_sheet(self, excel_file):
        data = []
        headers = list(self.mapping.keys())
        
        # Prepare data for writing
        for linie in self.linii:
            row = []
            for header in headers:
                mapping = self.mapping.get(header)
                value = self.helper.resolve_mapping(linie, mapping)
                value = "" if value in config.NULL_VALUES else value
                row.append(value)
            data.append(row)
        
        workbook = load_workbook(excel_file)
        sheet = workbook["LINIE_JOASA_TENSIUNE"]
        
        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}
        
        # Write data to the sheet
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if self.helper.n(header) in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value if cell_value is not None else "")
        
        workbook.save(excel_file)
