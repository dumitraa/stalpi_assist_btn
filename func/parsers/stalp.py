from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from qgis.core import QgsMessageLog, Qgis # type: ignore

class StalpJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, id_linie_jt_1, nr_crt_linie_jt_1, id_linie_jt_2, nr_crt_linie_jt_2, id_linie_jt_3, nr_crt_linie_jt_3, id_linie_jt_4, nr_crt_linie_jt_4, id_linie_jt_5, nr_crt_linie_jt_5, id_linie_jt_6, nr_crt_linie_jt_6, id_linie_jt_7, nr_crt_linie_jt_7, denum, nr_ins_stp, desc_det, prop, det_prop, tip_zona_amp, jud, prim, loc, tip_str, str, tip_cir, tip_mat, desc_ctg_mt_jt, nr_cir, uzura_stp, tip_fund, obs_fund, anc, obs_anc, adaos, obs_adaos, fib_opt, nr_cir_fo, prop_fo, ltc, nr_cir_ltc, prop_ltc, catv, nr_cir_catv, prop_catv, echip_com, disp_cuib_pas, nr_cons_c2s, nr_cons_c4s, nr_cons_c2t, nr_cons_c4t, nr_cons_c2br, nr_cons_c4br, tip_leg_jt, priza_leg_pam, corp_il, cutie_sel, geo, lat, long, alt, x_stereo_70, y_stereo_70, z_stereo_70, sursa_coord, data_coord, obs, img_file_1, img_file_2, img_file_3, img_file_4):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.id_linie_jt_1 = id_linie_jt_1
        self.nr_crt_linie_jt_1 = nr_crt_linie_jt_1
        self.id_linie_jt_2 = id_linie_jt_2
        self.nr_crt_linie_jt_2 = nr_crt_linie_jt_2
        self.id_linie_jt_3 = id_linie_jt_3
        self.nr_crt_linie_jt_3 = nr_crt_linie_jt_3
        self.id_linie_jt_4 = id_linie_jt_4
        self.nr_crt_linie_jt_4 = nr_crt_linie_jt_4
        self.id_linie_jt_5 = id_linie_jt_5
        self.nr_crt_linie_jt_5 = nr_crt_linie_jt_5
        self.id_linie_jt_6 = id_linie_jt_6
        self.nr_crt_linie_jt_6 = nr_crt_linie_jt_6
        self.id_linie_jt_7 = id_linie_jt_7
        self.nr_crt_linie_jt_7 = nr_crt_linie_jt_7
        self.denum = denum
        self.nr_ins_stp = nr_ins_stp
        self.desc_det = desc_det
        self.prop = prop
        self.det_prop = det_prop
        self.tip_zona_amp = tip_zona_amp
        self.jud = jud
        self.prim = prim
        self.loc = loc
        self.tip_str = tip_str
        self.str = str
        self.tip_cir = tip_cir
        self.tip_mat = tip_mat
        self.desc_ctg_mt_jt = desc_ctg_mt_jt
        self.nr_cir = nr_cir
        self.uzura_stp = uzura_stp
        self.tip_fund = tip_fund
        self.obs_fund = obs_fund
        self.anc = anc
        self.obs_anc = obs_anc
        self.adaos = adaos
        self.obs_adaos = obs_adaos
        self.fib_opt = fib_opt
        self.nr_cir_fo = nr_cir_fo
        self.prop_fo = prop_fo
        self.ltc = ltc
        self.nr_cir_ltc = nr_cir_ltc
        self.prop_ltc = prop_ltc
        self.catv = catv
        self.nr_cir_catv = nr_cir_catv
        self.prop_catv = prop_catv
        self.echip_com = echip_com
        self.disp_cuib_pas = disp_cuib_pas
        self.nr_cons_c2s = nr_cons_c2s
        self.nr_cons_c4s = nr_cons_c4s
        self.nr_cons_c2t = nr_cons_c2t
        self.nr_cons_c4t = nr_cons_c4t
        self.nr_cons_c2br = nr_cons_c2br
        self.nr_cons_c4br = nr_cons_c4br
        self.tip_leg_jt = tip_leg_jt
        self.priza_leg_pam = priza_leg_pam
        self.corp_il = corp_il
        self.cutie_sel = cutie_sel
        self.geo = geo
        self.lat = lat
        self.long = long
        self.alt = alt
        self.x_stereo_70 = x_stereo_70
        self.y_stereo_70 = y_stereo_70
        self.z_stereo_70 = z_stereo_70
        self.sursa_coord = sursa_coord
        self.data_coord = data_coord
        self.obs = obs
        self.img_file_1 = img_file_1
        self.img_file_2 = img_file_2
        self.img_file_3 = img_file_3
        self.img_file_4 = img_file_4

    def __repr__(self):
        return f"StalpJT(denum={self.denum}, niv_ten={self.niv_ten}, tip_lin={self.tip_lin})"


class IgeaStalpParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.stalpi: List[StalpJT] = []

        self.mapping = {
            "Nr crt": "nr_crt",
            "ID_linie JT": "id_linie_jt",
            "Denumire": "denum",
            "Descrierea BDI": ("STP.", " ", "denum", "str", ),
            "Numar inscriptionat pe stalp": "nr_ins_stp",
            "Descriere detaliata": "desc_det",
            "Proprietar": "prop",
            "Detaliere Proprietar": "det_prop",
            "Tip zona de amplasare": lambda stalp: stalp.tip_zona_amp if stalp.tip_zona_amp else "Rural",
            "Judet": "jud",
            "Primarie": "prim",
            "Localitate": "loc",
            "Tip strada": "tip_str",
            "Strada": "str",
            "Tip circuit": "tip_cir",
            "Tip material": "tip_mat",
            "Descriere catalog MT, JT": "desc_ctg_mt_jt",
            "Numar circuite": "nr_cir",
            "Defecte stalp": "", # ?
            "Tipul fundatiei": "tip_fund",
            "Observatii fundatie": "obs_fund",
            "Ancora": "anc",
            "Observatii ancora": "obs_anc",
            "Adaos": "adaos",
            "Observatii adaos": "obs_adaos",
            "Fibra optica": "fib_opt",
            "Numar circuite FO": "nr_cir_fo",
            "Proprietar FO": "prop_fo",
            "LTC": "ltc",
            "Numar circuite LTC": "nr_cir_ltc",
            "Proprietar LTC": "prop_ltc",
            "CATV": "catv",
            "Numar circuite CATV": "nr_cir_catv",
            "Proprietar CATV": "prop_catv",
            "Echipamente comunicatii": "echip_com",
            "Dispozitiv cuib pasari": "disp_cuib_pas",
            "Tipul de consola": "0", # what?
            "Tip legaturi JT": "tip_leg_jt",
            "Priza de legare la pamant": "priza_leg_pam",
            "Corp iluminat": "corp_il",
            "Cutie selectivitate/cutie sectionare": "cutie_sel",
            "Latitudine (grade zecimale)": lambda stalp: f"{stalp.lat:.8f}" if stalp.lat else "",
            "Longitudine (grade zecimale)": "long",
            "Altitudine (m)": "alt",
            "x - STEREO 70 (m)": "x_stereo_70",
            "y - STEREO 70 (m)": "y_stereo_70",
            "z - STEREO 70 (m)": "z_stereo_70",
            "Geometrie": "geo",
            "Sursa coordonate": "sursa_coord",
            "Data actualizarii coordonatelor": "data_coord",
            "Observatii": "obs"
        }
        
        # self.qgis_mapping = ["CLASS_ID", "ID_BDI", "NR_CRT", "ID_LINIE_JT", "NR_CRT_LINIE_JT_1", "ID_LINIE_JT_2", "NR_CRT_LINIE_JT_2", "ID_LINIE_JT_3", "NR_CRT_LINIE_JT_3", "ID_LINIE_JT_4", "NR_CRT_LINIE_JT_4", "ID_LINIE_JT_5", "NR_CRT_LINIE_JT_5", "ID_LINIE_JT_6", "NR_CRT_LINIE_JT_6", "ID_LINIE_JT_7", "NR_CRT_LINIE_JT_7", "DENUM", "NR_INS_STP", "DESC_DET", "PROP", "DET_PROP", "TIP_ZONA_AMP", "JUD", "PRIM", "LOC", "TIP_STR", "STR", "TIP_CIR", "TIP_MAT", "DESC_CTG_MT_JT", "NR_CIR", "UZURA_STP", "TIP_FUND", "OBS_FUND", "ANC", "OBS_ANC", "ADAOS", "OBS_ADAOS", "FIB_OPT", "NR_CIR_FO", "PROP_FO", "LTC", "NR_CIR_LTC", "PROP_LTC", "CATV", "NR_CIR_CAT", "PROP_CATV", "ECHIP_COM", "DISP_CUIB_PAS", "NR_CONS_C2S", "NR_CONS_C4S", "NR_CONS_C2T", "NR_CONS_C4T", "NR_CONS_C2BR", "NR_CONS_C4BR", "TIP_LEG_JT", "PRIZA_LEG_PAM", "CORP_IL", "CUTIE_SEL", "GEO", "LAT", "LONG", "ALT", "X_STEREO_70", "Y_STEREO_70", "Z_STEREO_70", "SURSA_COORD", "DATA_COORD", "OBS", "IMG_FILE_1", "IMG_FILE_2", "IMG_FILE_3", "IMG_FILE_4"]
            
            
    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        for feature in self.vector_layer.getFeatures():
            stalp_data = StalpJT(
                id=feature.id(),
                class_id=feature["CLASS_ID"],
                id_bdi=feature["ID_BDI"],
                nr_crt=feature["NR_CRT"],
                id_linie_jt_1=feature["ID_LINIE_JT_1"],
                nr_crt_linie_jt_1=feature["NR_CRT_LINIE_JT_1"],
                id_linie_jt_2=feature["ID_LINIE_JT_2"],
                nr_crt_linie_jt_2=feature["NR_CRT_LINIE_JT_2"],
                id_linie_jt_3=feature["ID_LINIE_JT_3"],
                nr_crt_linie_jt_3=feature["NR_CRT_LINIE_JT_3"],
                id_linie_jt_4=feature["ID_LINIE_JT_4"],
                nr_crt_linie_jt_4=feature["NR_CRT_LINIE_JT_4"],
                id_linie_jt_5=feature["ID_LINIE_JT_5"],
                nr_crt_linie_jt_5=feature["NR_CRT_LINIE_JT_5"],
                id_linie_jt_6=feature["ID_LINIE_JT_6"],
                nr_crt_linie_jt_6=feature["NR_CRT_LINIE_JT_6"],
                id_linie_jt_7=feature["ID_LINIE_JT_7"],
                nr_crt_linie_jt_7=feature["NR_CRT_LINIE_JT_7"],
                denum=feature["DENUM"],
                nr_ins_stp=feature["NR_INS_STP"],
                desc_det=feature["DESC_DET"],
                prop=feature["PROP"],
                det_prop=feature["DET_PROP"],
                tip_zona_amp=feature["TIP_ZONA_AMP"],
                jud=feature["JUD"],
                prim=feature["PRIM"],
                loc=feature["LOC"],
                tip_str=feature["TIP_STR"],
                str=feature["STR"],
                tip_cir=feature["TIP_CIR"],
                tip_mat=feature["TIP_MAT"],
                desc_ctg_mt_jt=feature["DESC_CTG_MT_JT"],
                nr_cir=feature["NR_CIR"],
                uzura_stp=feature["UZURA_STP"],
                tip_fund=feature["TIP_FUND"],
                obs_fund=feature["OBS_FUND"],
                anc=feature["ANC"],
                obs_anc=feature["OBS_ANC"],
                adaos=feature["ADAOS"],
                obs_adaos=feature["OBS_ADAOS"],
                fib_opt=feature["FIB_OPT"],
                nr_cir_fo=feature["NR_CIR_FO"],
                prop_fo=feature["PROP_FO"],
                ltc=feature["LTC"],
                nr_cir_ltc=feature["NR_CIR_LTC"],
                prop_ltc=feature["PROP_LTC"],
                catv=feature["CATV"],
                nr_cir_catv=feature["NR_CIR_CATV"],
                prop_catv=feature["PROP_CATV"],
                echip_com=feature["ECHIP_COM"],
                disp_cuib_pas=feature["DISP_CUIB_PAS"],
                nr_cons_c2s=feature["NR_CONS_C2S"],
                nr_cons_c4s=feature["NR_CONS_C4S"],
                nr_cons_c2t=feature["NR_CONS_C2T"],
                nr_cons_c4t=feature["NR_CONS_C4T"],
                nr_cons_c2br=feature["NR_CONS_C2BR"],
                nr_cons_c4br=feature["NR_CONS_C4BR"],
                tip_leg_jt=feature["TIP_LEG_JT"],
                priza_leg_pam=feature["PRIZA_LEG_PAM"],
                corp_il=feature["CORP_IL"],
                cutie_sel=feature["CUTIE_SEL"],
                geo=feature["GEO"],
                lat=feature["LAT"],
                long=feature["LONG"],
                alt=feature["ALT"],
                x_stereo_70=feature["X_STEREO_70"],
                y_stereo_70=feature["Y_STEREO_70"],
                z_stereo_70=feature["Z_STEREO_70"],
                sursa_coord=feature["SURSA_COORD"],
                data_coord=feature["DATA_COORD"],
                obs=feature["OBS"],
                img_file_1=feature["IMG_FILE_1"],
                img_file_2=feature["IMG_FILE_2"],
                img_file_3=feature["IMG_FILE_3"],
                img_file_4=feature["IMG_FILE_4"]
            )
            self.stalpi.append(stalp_data)
            
    def get_name(self):
        return "STALP_XML_"
            
    def get_data(self):
        return self.stalpi

    def resolve_mapping(self, parser, mapping):
        if isinstance(mapping, tuple):
            parts = [
                getattr(parser, element, "").strip() if hasattr(parser, element) else str(element).strip()
                for element in mapping
            ]
            return " ".join(filter(None, parts)).strip()
        elif callable(mapping):
            # If mapping is a function, execute it
            return mapping(parser)
        return getattr(parser, mapping, "") if mapping else ""
    
    def write_to_excel_sheet(self, excel_file):
        QgsMessageLog.logMessage(f"Writing stalpi to excel file: {excel_file}", "StalpiAssist", level=Qgis.Info)
        data = []
        headers = list(self.mapping.keys())
        
        # Prepare data for writing
        for stalp in self.stalpi:
            row = []
            for header in headers:
                mapping = self.mapping.get(header)
                value = self.resolve_mapping(stalp, mapping)
                value = "" if value in ["NULL", None, "nan"] else value
                row.append(value)
                QgsMessageLog.logMessage(f"Writing {value} to row with header {header}", "StalpiAssist", level=Qgis.Info)
            data.append(row)
        QgsMessageLog.logMessage(f"Prepared data with {len(data)} rows", "StalpiAssist", level=Qgis.Info)
        
        workbook = load_workbook(excel_file)
        sheet = workbook["STÂLP"]
        
        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}
        QgsMessageLog.logMessage(f"Existing headers found: {existing_headers}", "StalpiAssist", level=Qgis.Info)
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if header.strip() in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header.strip()], value=cell_value if cell_value is not None else "")
        QgsMessageLog.logMessage(f"Data written to the sheet starting from row {start_row}", "StalpiAssist", level=Qgis.Info)
        
        # Add borders to the cells
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for header in headers:
                if header.strip() in existing_headers:
                    cell = sheet.cell(row=row_idx, column=existing_headers[header.strip()])
                    cell.border = thin_border
        QgsMessageLog.logMessage(f"Borders added to cells for rows {start_row} to {start_row + len(data) - 1}", "StalpiAssist", level=Qgis.Info)
        
        workbook.save(excel_file)
        QgsMessageLog.logMessage(f"Excel file {excel_file} saved successfully", "StalpiAssist", level=Qgis.Info)
