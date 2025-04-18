from typing import List
from openpyxl import load_workbook
from ... import config
from ..helper_functions import HelperBase


class GrupMasuraJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, denum, class_id_loc, id_loc, nr_crt_loc, class_id_inst_sup, id_inst_sup, nr_crt_inst_sup, jud, prim, loc, tip_str, str, nr_scara, etaj, ap):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.denum = denum
        self.class_id_loc = class_id_loc
        self.id_loc = id_loc
        self.nr_crt_loc = nr_crt_loc
        self.class_id_inst_sup = class_id_inst_sup
        self.id_inst_sup = id_inst_sup
        self.nr_crt_inst_sup = nr_crt_inst_sup
        self.jud = jud
        self.prim = prim
        self.loc = loc
        self.tip_str = tip_str
        self.str = str
        self.nr_scara = nr_scara
        self.etaj = etaj
        self.ap = ap

    def __repr__(self):
        return f"GrupMasuraJT(denum={self.denum})"


class IgeaGrupMasuraParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.grupuri: List[GrupMasuraJT] = []
        self.helper = HelperBase()

        self.mapping = {
            "Nr.crt": "nr_crt",
            "Denumire": lambda gr: gr.denum['correct'],
            "Descrierea BDI": lambda gr: f"GRUP MASURA {gr.denum['correct']}",
            "Nr.crt_Locatia": "nr_crt_loc",
            "Locatia": lambda gr: f"FB {gr.denum['short_denum']}",
            "ID_Descrierea instalatiei uperioare": "class_id_inst_sup",
            "Descrierea instalatiei uperioare": "nr_crt_inst_sup",
            "Judet": "jud",
            "Primarie": "prim",
            "Localitate": "loc",
            "Tip strada": "tip_str",
            "Strada": lambda gr: gr.denum['str'],
            "nr./ scara": lambda gr: gr.denum['nr'],
            "Etaj": "etaj",
            "Apartament": "ap"
        }
        
        # self.qgis_mapping = ["CLASS_ID", "ID_BDI", "NR_CRT", "DENUM", "CLASS_ID_LOC", "ID_LOC", "NR_CRT_LOC", "CLASS_ID_INST_SUP", "ID_INST_SUP", "NR_CRT_INST_SUP", "JUD", "PRIM", "LOC", "TIP_STR", "STR", "NR_SCARA", "ETAJ", "AP"]
            
    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        features = list(self.vector_layer.getFeatures())
        for feature in features:
            correct_denum = self.helper.get_correct_denum(feature)
            attributes = {key: feature[key] for key in feature.fields().names()}
            grup_data = GrupMasuraJT(
                id=feature.id(),
                class_id=attributes.get("CLASS_ID"),
                id_bdi=attributes.get("ID_BDI"),
                nr_crt=attributes.get("NR_CRT"),
                denum={'initial': attributes.get("DENUM"),
                       'correct': correct_denum['denum'],
                       'short_denum': correct_denum['short_denum'],
                       'nr': correct_denum['nr_scara'],
                       'str': correct_denum['str']
                       }, 
                class_id_loc=attributes.get("CLASS_ID_LOC"),
                id_loc=attributes.get("ID_LOC"),
                nr_crt_loc=attributes.get("NR_CRT_LOC"),
                class_id_inst_sup=attributes.get("CLASS_ID_LOC"),
                id_inst_sup=attributes.get("ID_INST_SUP"),
                nr_crt_inst_sup=attributes.get("NR_CRT_INST_SUP"),
                jud=attributes.get("JUD"),
                prim=attributes.get("PRIM"),
                loc=attributes.get("LOC"),
                tip_str=attributes.get("TIP_STR"),
                str=attributes.get("STR"),
                nr_scara=attributes.get("NR_SCARA"),
                etaj=attributes.get("ETAJ"),
                ap=attributes.get("AP")
            )
            self.grupuri.append(grup_data)
    
    def get_name(self):
        return "GRUP_MASURA_XML_"
    
    def get_data(self):
        return self.grupuri

    def write_to_excel_sheet(self, excel_file):
        data = []
        headers = list(self.mapping.keys())
        
        # Prepare data for writing
        sorted_gr = sorted(
            self.grupuri,
            key=lambda gr: gr.nr_crt if gr.nr_crt not in config.NULL_VALUES else float("inf")
        )
        for grupa in sorted_gr:
            row = []
            for header in headers:
                mapping = self.mapping.get(header)
                value = self.helper.resolve_mapping(grupa, mapping)
                # Replace None with an empty string
                value = "" if value in config.NULL_VALUES else value
                row.append(value)
            data.append(row)
        
        workbook = load_workbook(excel_file)
        sheet = workbook["GRUP_MASURA"]
        
        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}
        
        # Write data to the sheet
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if self.helper.n(header) in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value if cell_value is not None else "")

        workbook.save(excel_file)

