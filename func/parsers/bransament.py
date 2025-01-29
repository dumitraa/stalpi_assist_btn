from typing import List
from openpyxl import load_workbook
from qgis.core import QgsMessageLog, Qgis, QgsProject  # type: ignore

class BransamentJT():
    def __init__(self, id, id_bdi, nr_crt, denum, class_id_loc, id_loc, nr_crt_loc, 
                 class_id_plc_br, id_plc_br, nr_crt_plc_br, tip_br, tip_cond, lung, jud, 
                 prim, loc, tip_str, street, nr_imob, geo, sursa_coord, data_coord, obs, geometry):
        self.id = id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.denum = denum
        self.class_id_loc = class_id_loc
        self.id_loc = id_loc
        self.nr_crt_loc = nr_crt_loc
        self.class_id_plc_br = class_id_plc_br
        self.id_plc_br = id_plc_br
        self.nr_crt_plc_br = nr_crt_plc_br
        self.tip_br = tip_br
        self.tip_cond = tip_cond
        self.lung = lung
        self.jud = jud
        self.prim = prim
        self.loc = loc
        self.tip_str = tip_str
        self.street = street
        self.nr_imob = nr_imob
        self.geo = geo
        self.sursa_coord = sursa_coord
        self.data_coord = data_coord
        self.obs = obs
        self.geometry = geometry

    def get_geometry(self):
        return self.geometry

    def __repr__(self):
        return f"BransamentJT(nr_crt={self.nr_crt}, denum={self.denum}, geo={self.geo})"

class IgeaBransamentParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.bransamente: List[BransamentJT] = []
        self.mapping = {
            "Nr.crt": "nr_crt",
            "Denumire": "denum",
            "Descrierea BDI": ("BR ", "denum"),
            "ID_Locatia": "id_loc",
            "Locatia": lambda br: self.get_linie_value(br),
            "ID_PAPT/Nr. Crt_Plecare bransament": "nr_crt_plc_br",
            "Plecare bransament": lambda br: self.get_stalpi_value(br),
            "Tip bransament": "tip_br",
            "Tipul dispunerii": lambda br: "LES" if "XABY" in br.tip_cond or "ACYABY" in br.tip_cond else "LEA",
            "Tip conductor": "tip_cond",
            "Lungime (m)": "lung",
            "Judet": "jud",
            "Primarie": "prim",
            "Localitate": "loc",
            "Tip strada": "tip_str",
            "Strada": "street",
            "Numar imobil": "nr_imob",
            "Geometrie": "geo",
            "Sursa coordonate": "sursa_coord",
            "Data actualizarii coordonatelor": "data_coord",
            "Observatii": "obs",
        }

    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        features = list(self.vector_layer.getFeatures())
        for feature in features:
            bransament_data = BransamentJT(
                id=feature.id(),
                id_bdi=feature['ID_BDI'],
                nr_crt=feature['NR_CRT'],
                denum=feature['DENUM'],
                class_id_loc=feature['CLASS_ID_LOC'],
                id_loc=feature['ID_LOC'],
                nr_crt_loc=feature['NR_CRT_LOC'],
                class_id_plc_br=feature['CLASS_ID_PLC_BR'],
                id_plc_br=feature['ID_PLC_BR'],
                nr_crt_plc_br=feature['NR_CRT_PLC_BR'],
                tip_br=feature['TIP_BR'],
                tip_cond=feature['TIP_COND'],
                lung=feature['LUNG'],
                jud=feature['JUD'],
                prim=feature['PRIM'],
                loc=feature['LOC'],
                tip_str=feature['TIP_STR'],
                street=feature['STR'],
                nr_imob=feature['NR_IMOB'],
                geo=feature['GEO'],
                sursa_coord=feature['SURSA_COORD'],
                data_coord=feature['DATA_COORD'],
                obs=feature['OBS'],
                geometry=feature.geometry()  # Pass the feature's geometry here
            )
            self.bransamente.append(bransament_data)

    def get_name(self):
        return "BRANSAMENT_XML_"

    def get_data(self):
        return self.bransamente

    def resolve_mapping(self, parser, mapping):
        if isinstance(mapping, tuple):
            parts = [
                str(getattr(parser, element, "")).strip() if hasattr(parser, element) else str(element).strip()
                for element in mapping
            ]
            return " ".join(filter(None, parts)).strip()
        elif callable(mapping):
            return mapping(parser)
        return str(getattr(parser, mapping, "")).strip() if mapping else ""

    def write_to_excel_sheet(self, excel_file):
        data = []
        headers = list(self.mapping.keys())
        sorted_br = sorted(
            self.bransamente,
            key=lambda br: br.nr_crt if br.nr_crt not in [None, "NULL", "nan"] else float("inf")
        )
        for bransament in sorted_br:
            row = []
            for header in headers:
                mapping = self.mapping[header]
                value = self.resolve_mapping(bransament, mapping)
                value = "" if value in ["NULL", None, "nan"] else value
                row.append(value)
            data.append(row)

        try:
            workbook = load_workbook(excel_file)
            sheet = workbook["BRANSAMENT"]
        except Exception as e:
            return

        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}

        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if header.strip(" ") in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value)

        try:
            workbook.save(excel_file)
        except Exception as e:
            QgsMessageLog.logMessage(f"Error saving workbook: {e}", "StalpiAssist", level=Qgis.Critical)

    def get_stalpi_value(self, bransament_feature):
        if not bransament_feature.get_geometry():
            QgsMessageLog.logMessage("Bransament feature has no geometry.", "StalpiAssist", level=Qgis.Warning)
            return ""

        stalp_layer = QgsProject.instance().mapLayersByName('STALP_XML_')[0]
        if not stalp_layer:
            QgsMessageLog.logMessage("STALP_XML_ layer not found.", "StalpiAssist", level=Qgis.Critical)
            return ""

        stalpi_features = stalp_layer.getFeatures()
        intersecting_features = [
            stalp for stalp in stalpi_features 
            if stalp.geometry() and stalp.geometry().intersects(bransament_feature.get_geometry())
        ]

        if not intersecting_features:
            QgsMessageLog.logMessage("No intersecting features found.", "StalpiAssist", level=Qgis.Warning)
            return ""

        aggregated_value = [
            f"STP. {stalp['DENUM']} {stalp['STR']}" for stalp in intersecting_features
        ]
        return aggregated_value[0] if aggregated_value else ""

    def get_linie_value(self, feature):
        '''
        Match bransament_feature's id_loc with LINIE_JT's ID_BDI and return LINIE_JT's DENUM.
        '''
        linie_layer = QgsProject.instance().mapLayersByName('LINIE_JT')[0]
        if not linie_layer:
            QgsMessageLog.logMessage("LINIE_JT layer not found.", "StalpiAssist", level=Qgis.Critical)
            return ""
        
        linie_features = linie_layer.getFeatures()
        matching_feature = [
            linie for linie in linie_features
            if linie['ID_BDI'] == feature.id_loc  # Use bransament_feature's id_loc attribute
        ]
        
        if not matching_feature:
            QgsMessageLog.logMessage("No matching feature found.", "StalpiAssist", level=Qgis.Warning)
            return ""
        
        return matching_feature[0]['DENUM'] if matching_feature else ""
