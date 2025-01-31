from typing import List
from openpyxl import load_workbook
from qgis.core import QgsMessageLog, Qgis, QgsProject, QgsFeatureRequest # type: ignore
from ..helper_functions import HelperBase


class DeschidereJT:
    def __init__(self, id, class_id, id_bdi, nr_crt, denum, ip_stp_inc, nr_crt_stp_inc, id_stp_term, nr_crt_stp_term, id_tr_jt1, nr_crt_tr_jt1, id_tr_jt2, nr_crt_tr_jt2, id_tr_jt3, nr_crt_tr_jt3, id_tr_jt4, nr_crt_tr_jt4, id_tr_jt5, nr_crt_tr_jt5, id_tr_jt6, nr_crt_tr_jt6, geo, lung, sursa_coord, data_coord, id_loc, locatia):
        self.id = id
        self.class_id = class_id
        self.id_bdi = id_bdi
        self.nr_crt = nr_crt
        self.denum = denum
        self.ip_stp_inc = ip_stp_inc
        self.nr_crt_stp_inc = nr_crt_stp_inc
        self.id_stp_term = id_stp_term
        self.nr_crt_stp_term = nr_crt_stp_term
        self.id_tr_jt1 = id_tr_jt1
        self.nr_crt_tr_jt1 = nr_crt_tr_jt1
        self.id_tr_jt2 = id_tr_jt2
        self.nr_crt_tr_jt2 = nr_crt_tr_jt2
        self.id_tr_jt3 = id_tr_jt3
        self.nr_crt_tr_jt3 = nr_crt_tr_jt3
        self.id_tr_jt4 = id_tr_jt4
        self.nr_crt_tr_jt4 = nr_crt_tr_jt4
        self.id_tr_jt5 = id_tr_jt5
        self.nr_crt_tr_jt5 = nr_crt_tr_jt5
        self.id_tr_jt6 = id_tr_jt6
        self.nr_crt_tr_jt6 = nr_crt_tr_jt6
        self.geo = geo
        self.lung = lung
        self.sursa_coord = sursa_coord
        self.data_coord = data_coord
        self.id_loc = id_loc
        self.locatia = locatia
        

    def __repr__(self):
        return f"DeschidereJT(denum={self.denum}, niv_ten={self.niv_ten}, tip_lin={self.tip_lin})"


class IgeaDeschidereParser:
    def __init__(self, vector_layer):
        self.vector_layer = vector_layer
        self.deschideri: List[DeschidereJT] = []
        self.helper = HelperBase()
        
        self.layers = self.get_layers()
        if not self.layers:
            QgsMessageLog.logMessage("No layers found matching the required names.", "StalpiAssist", level=Qgis.Critical)
            return

        self.mapping = {
            "Nr.crt": "nr_crt",
            "Denumire": "denum",
            "Descrierea BDI": ("DESC", "denum"),
            "ID_Locatia": lambda ds: ds.id_loc,
            "Locatia": lambda ds: ds.locatia,
            "Nr.crt_Inceput": "nr_crt_stp_inc",
            "Stâlpul de inceput": lambda ds: ds.denum.split('-')[0].strip() if ds.denum else "",
            "Nr.crt_sfarsit": "nr_crt_stp_term",
            "Stâlpul terminal": lambda ds: ds.denum.split('-')[1].strip() if ds.denum else "",
            "ID_Tronson JT1": "id_tr_jt1",
            "Tronson JT1": "nr_crt_tr_jt1",
            "ID_Tronson JT2": "id_tr_jt2",
            "Tronson JT2": "nr_crt_tr_jt2",
            "ID_Tronson JT3": "id_tr_jt3",
            "Tronson JT3": "nr_crt_tr_jt3",
            "ID_Tronson JT4": "id_tr_jt4",
            "Tronson JT4": "nr_crt_tr_jt4",
            "ID_Tronson JT5": "id_tr_jt5",
            "Tronson JT5": "nr_crt_tr_jt5",
            "ID_Tronson JT6": "id_tr_jt6",
            "Tronson JT6": "nr_crt_tr_jt6",
            "Lungime (m)": "lung",
            "Geometrie": "geo",
            "Observatii": lambda ds: ""
        }
        
        # self.qgis_mapping = ["CLASS_ID", "ID_BDI", "NR_CRT", "DENUM", "ID_STP_INC", "NR_CRT_STP_INC", "ID_STP_TERM", "NR_CRT_STP_TERM", "ID_TR_JT1", "NR_CRT_TR_JT1", "ID_TR_JT2", "NR_CRT_TR_JT2", "ID_TR_JT3", "NR_CRT_TR_JT3", "ID_TR_JT4", "NR_CRT_TR_JT4", "ID_TR_JT5", "NR_CRT_TR_JT5", "ID_TR_JT6", "NR_CRT_TR_JT6", "GEO", "LUNG", "SURSA_COORD", "DATA_COORD"]
            
    def parse(self):
        if not self.vector_layer.isValid():
            raise ValueError("The provided layer is not valid.")

        features = list(self.vector_layer.getFeatures())
        for feature in features:
            try:
                loc_data = self.get_location_data(feature["NR_CRT"], self.layers.get("DESCHIDERI MACHETA"))
            except Exception as e:
                QgsMessageLog.logMessage(f"Error getting location data: {e}", "StalpiAssist", level=Qgis.Critical)
                loc_data = {"id_loc": "", "locatia": ""}
            
            # Cache loc_data for mapping
            feature_loc_data = {
                "id_loc": loc_data["id_loc"] if loc_data["id_loc"] else "",
                "locatia": loc_data["locatia"] if loc_data["locatia"] else ""
            }
            
            deschidere_data = DeschidereJT(
                id=feature.id(),
                class_id=feature["CLASS_ID"],
                id_bdi=feature["ID_BDI"],
                nr_crt=feature["NR_CRT"],
                denum=feature["DENUM"],
                ip_stp_inc=feature["ID_STP_INC"],
                nr_crt_stp_inc=feature["NR_CRT_STP_INC"],
                id_stp_term=feature["ID_STP_TERM"],
                nr_crt_stp_term=feature["NR_CRT_STP_TERM"],
                id_tr_jt1=feature["ID_TR_JT1"],
                nr_crt_tr_jt1=feature["NR_CRT_TR_JT1"],
                id_tr_jt2=feature["ID_TR_JT2"],
                nr_crt_tr_jt2=feature["NR_CRT_TR_JT2"],
                id_tr_jt3=feature["ID_TR_JT3"],
                nr_crt_tr_jt3=feature["NR_CRT_TR_JT3"],
                id_tr_jt4=feature["ID_TR_JT4"],
                nr_crt_tr_jt4=feature["NR_CRT_TR_JT4"],
                id_tr_jt5=feature["ID_TR_JT5"],
                nr_crt_tr_jt5=feature["NR_CRT_TR_JT5"],
                id_tr_jt6=feature["ID_TR_JT6"],
                nr_crt_tr_jt6=feature["NR_CRT_TR_JT6"],
                geo=feature["GEO"],
                lung=feature["LUNG"],
                sursa_coord=feature["SURSA_COORD"],
                data_coord=feature["DATA_COORD"],
                id_loc = feature_loc_data["id_loc"],
                locatia = feature_loc_data["locatia"]
            )

            self.deschideri.append(deschidere_data)
    
    def get_name(self):
        return "DESCHIDERI_XML_"
            
    def get_data(self):
        return self.deschideri
    
    def resolve_mapping(self, parser, mapping):
        if isinstance(mapping, tuple):
            parts = [
                str(getattr(parser, element, "")).strip() if hasattr(parser, element) else str(element).strip()
                for element in mapping
            ]
            return " ".join(filter(None, parts)).strip()
        elif callable(mapping):
            return mapping(parser)
        return str(getattr(parser, mapping, "")).strip() if mapping else ""


    def write_to_excel_sheet(self, excel_file):        
        data = []
        headers = list(self.mapping.keys())
        
        sorted_ds = sorted(
            self.deschideri,
            key=lambda ds: ds.nr_crt if ds.nr_crt not in [None, "None", "NULL", "nan"] else float("inf")
        )
        for deschidere in sorted_ds:
            row = []
            for header in headers:
                mapping = self.mapping.get(header)
                value = self.resolve_mapping(deschidere, mapping)
                value = "" if value in ["NULL", "None", None, "nan"] else value
                row.append(value)
            data.append(row)
        
        workbook = load_workbook(excel_file)
        sheet = workbook["DESCHIDERE"]
        
        start_row = sheet.max_row + 1
        header_row = sheet.max_row - 1
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}
        
        # Write data to the sheet
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if header.strip(" ") in existing_headers:
                    sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value)
        
        workbook.save(excel_file)


    def get_location_data(self, fid, layer):
        """
        Retrieves location data from the 'DESCHIDERI MACHETA' layer.
        
        :param fid: Feature ID to search for
        :param layer: The layer to search within
        :return: Dictionary with "id_loc" and "locatia" fields
        """
        
        if not layer:
            QgsMessageLog.logMessage(f"Layer 'DESCHIDERI MACHETA' not found!", "StalpiAssist", level=Qgis.Critical)
            return {"id_loc": "", "locatia": ""}

        # Use QgsFeatureRequest to filter by fid
        request = QgsFeatureRequest().setFilterExpression(f'"Nr.crt" = {fid}')

        feature = next(layer.getFeatures(request), None)

        if not feature:
            QgsMessageLog.logMessage(f"Feature with Nr.crt {fid} not found!", "StalpiAssist", level=Qgis.Critical)
            return {"id_loc": "", "locatia": ""}

        # Extract fields with fallback for None, "NULL", or "nan"
        def clean_value(value):
            return "" if value in [None, "None", "NULL", "nan"] else str(value)

        id_loc = clean_value(feature["ID_Locatia"])
        locatia = clean_value(feature["Locatia"])

        return {"id_loc": id_loc, "locatia": locatia}



    def get_layers(self):
        '''
        Get layers by name from the QGIS project and add them to self.layers
        '''
        layers = {}
        layer_names = ["TRONSON_predare_xml", "LINIE_JT", 'DESCHIDERI_XML_', 'DESCHIDERI MACHETA']

        # Get all layers in the current QGIS project (keep the layer objects)
        qgis_layers = QgsProject.instance().mapLayers().values()

        # Iterate through the actual layer objects
        for layer_name in layer_names:
            layer = next((l for l in qgis_layers if l.name() == layer_name), None)
            layers[layer_name] = layer  # Add the layer if found, else None

        return layers